from __future__ import absolute_import, division, \
    print_function, unicode_literals
import pandas as pd
import numpy as np
import openpyxl
import json


def show_func():
    print("+-------------------------------+")
    print("|model online tools             |")
    print("+-------------------------------+")
    print("|1.dump_feats_json              |")
    print("|2.clean_table                  |")
    print("|3.get_feats_map                |")
    print("+-------------------------------+")


def dump_feats_json(model_name, feat_names, feat_path):
    """Save the feature file as JSON.

    Parameters
    ----------
    model_name : str
        Model bin file name

    feat_names : list
        Feature list name

    feat_path : str
        The path to save features file


    Returns
    --------
    cutoff : float
        The threshold corresponding to KS value.

    Examples
    ----------
    >>> feat_path = "path/to/sample/xxx.json"
    >>> feat_names = ["aaa", "bbb", "ccc"]
    >>> model_name = "templete_model.bin"
    >>> dump_feats_json(model_name, feat_names, feat_path)

    Notes
    -----
    This method only deals with the case that the features are numerical.
    """

    feats_dict = {"ntree_limit": 0, "model_file_name": model_name}
    feats_array = [{"name": feat_name, "method": "DirectFloat"}
                   for feat_name in feat_names]
    feats_dict["features"] = feats_array
    with open(feat_path, "w", encoding="utf-8") as json_file:
        json.dump(feats_dict, json_file,
                  separators=(',', ': '),
                  indent=4, ensure_ascii=False)


def clean_table(intput_path, output_path):
    """Split the merged cells, and make sure that each row after
    splitting is the same as the value of the original cell.

    Parameters
    ----------
    intput_path : str
        Merge_cells excel path to be processed

    output_path : str
        Unmerge_cells excel path to be saved

    Returns
    --------
    None

    Examples
    ----------
    >>> intput_path = "path/to/sample/xxx.xlsx"
    >>> output_path = "path/to/sample/yyy.xlsx"
    >>> clean_table(intput_path, output_path)

    Notes
    -----
    This method only deals with the case that the features are numerical.
    """

    # data_only=True indicates that if the cell contains a calculation
    # formula, only the calculation result is read
    workbook = openpyxl.load_workbook(intput_path, data_only=True)
    name_list = workbook.sheetnames

    # Traverse all sheets of Excel
    for name in name_list:
        cursheet = workbook[name]

        # Get the location information of merged cells in
        # the current sheet (MultiCellRange is an iterative object)
        merge_list = cursheet.merged_cells

        # Save location information of merged cells
        cr = []
        for merge_area in merge_list:

            # Get the starting and ending row coordinates
            # and column coordinates of merged cells
            r1, r2, c1, c2 = merge_area.min_row, merge_area.max_row, \
                             merge_area.min_col, merge_area.max_col

            # If the difference of row coordinates is greater
            # than 0, it means that the cells
            # are merged and the coordinates are added
            if r2 - r1 > 0:
                cr.append((r1, r2, c1, c2))

        # split cell
        for r1, r2, c1, c2 in cr:

            cursheet.unmerge_cells(start_row=r1,
                                   end_row=r2, start_column=c1, end_column=c2)

            # Fill the remaining cells after splitting
            # with the information of the first cell
            for row in range(r1, r2):
                cursheet.cell(row=row + 1, column=c1,
                              value=cursheet.cell(r1, c1).value)

    workbook.save(output_path)


def get_feats_map(input_path, feat_names, wrong_sheets, output_path):
    """Save the profile schema of decision platform

    Parameters
    ----------
    input_path : str
        Excel path to traverse

    feat_names: list
        The hive feature name used in the model

    wrong_sheets : list
        Dirty sheets of excel not to be dealt with

    output_path : str
        The path to save Schema infos on decision platform

    Returns
    --------
    None

    Examples
    ----------
    >>> intput_path = "path/to/sample/yyy.xlsx"
    >>> feat_names = ["aaa", "bbb", "ccc"]
    >>> wrong_sheets = ["ddd", "eee", "fff"]
    >>> output_path = "path/to/sample/zzz.txt"
    >>> clean_table(intput_path, output_path)

    Notes
    -----
    This method only deals with the case that the features are numerical.
    """

    # Read all sheets in the financial portfile excel
    data_dict = pd.read_excel(input_path, sheet_name=None, index_col=None)
    res, marked = [("zero", "zero")] * len(feat_names), [0] * len(feat_names)

    # Add the porfile group name and the porfile feature name to res
    def find_group_feat(data_df):
        for index, row in data_df.iterrows():

            # Get the hive feature name, porfile group name and
            # porfile feature name of the current line
            hive_feature_name = row["hive字段名 / kafka 字段名"] \
                if row["hive字段名 / kafka 字段名"] is not np.nan else ""
            profile_group_name = row["组名"] if row["组名"] is not np.nan else ""
            profile_feature_name = row["画像特征名"] if \
                row["画像特征名"] is not np.nan else "xxx"

            # Traverse the hive feature name to match
            for i, colname in enumerate(feat_names):
                if marked[i] == 1:
                    continue
                if colname == hive_feature_name:
                    marked[i] = 1
                    res[i] = (profile_group_name, profile_feature_name)

    # Remove all unavailable sheets
    for key, tmp_df in data_dict.items():

        if "组名" not in tmp_df.columns or "hive字段名 / kafka 字段名" not in tmp_df.columns or "画像特征名" not in tmp_df.columns:
            print("current sheet: {} has incomplete information，skip current sheet".format(key))
            continue

        find_group_feat(tmp_df)

    unmatched, count = [], 0
    res_map_str = "map("
    for index, feat in enumerate(feat_names):
        if res[index][0] == "zero":
            unmatched.append(feat)
            count += 1
        else:
            assert len(res[index][0]) != 0 and len(res[index][1]) != 0, \
                "The length of profile group name " \
                "or profile feature name cannot be 0"
            cur_groupname = res[index][0].strip("\n")
            res_map_str = res_map_str + "\"%s\", MapGet($%s, \"%s\"), " % (
                feat, cur_groupname, res[index][1])

    assert count == 0, "There is a hive feature whose " \
                       "profile group name is not found"
    schema = res_map_str[:-2] + ")"

    # Save mapping results
    with open(output_path, "w") as f:
        f.write(schema)
